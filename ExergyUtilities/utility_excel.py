# ExergyUtilities
# Copyright (c) 2010, B. Marcus Jones <>
# All rights reserved.
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""The :mod:`xx` module is a utility and notebook module to interface with Excel. The two xlrd/xlwt  
"""

#raise Exception, "OBSELETE?>>?>"

from win32com.client import Dispatch
import logging.config
import os
import re
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
#from ExergyUtilities.utility_inspect import get_self, get_parent
from utility_inspect import get_self, get_parent
import unittest
from config import *

import shutil

class excel_book_write(object):
    def __init__(self):
        #self.excelPath = excelPath
        logging.debug("Excel file at: {}, exists={}".format(self.excelPath,self.exists))

    @property
    def exists(self):
        return os.path.exists(self.excelPath)

def excel_write_table_xlsx(fullPath,sheetName,rows):
    wb = Workbook()
    sheet = wb.create_sheet()
    sheet.title = sheetName


    for i,row in enumerate(rows):
        i = i + 1

        col = get_column_letter(i)
        for j,item in enumerate(row):
            j = j + 1
            sheet.cell("{}{}".format(col,j)).value = item

    #    :
    #        sheet.write(i, j, item)

    wb.save(fullPath)
            #sheet0, 0
    logging.debug("Wrote {} rows to {}".format(len(rows),fullPath))

            #print i,row
        #print enumerate(rows)



def excel_write_table(fullPath,sheetName,rows):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet(sheetName)
    print(sheet)

    for i,row in enumerate(rows):
        for j,item in enumerate(row):
            sheet.write(i, j, item)

    wb.save(fullPath)
            #sheet0, 0
    logging.debug("Wrote {} rows to {}".format(len(rows),fullPath))

            #print i,row
        #print enumerate(rows)

class ExcelBookRead2(object):
    """
    import xlrd
    book = xlrd.open_workbook("myfile.xls")
    print "The number of worksheets is", book.nsheets
    print "Worksheet name(s):", book.sheet_names()
    sh = book.sheet_by_index(0)
    print sh.name, sh.nrows, sh.ncols
    print "Cell D30 is", sh.cell_value(rowx=29, colx=3)
    for rx in range(sh.nrows):
        print sh.row(rx)
    # Refer to docs for more details.
    # Feedback on API is welcomed.
    """


    def __enter__(self):
        logging.debug("***Enter***".format())
        self.wb = xlrd.open_workbook(self.excelPath)
        return self

    def __exit__(self, type, value, traceback):
        logging.debug("***Exit***".format())
        
    def __init__(self, excelPath):
        self.excelPath = excelPath
        logging.debug("Excel file at: {}".format(self.excelPath))
        
        
    def get_sheet_names(self):

        sheetNames = self.wb.sheet_names()
        logging.debug("Found {} sheet names: {}".format(len(sheetNames),sheetNames))
        return sheetNames

    def get_table_literal(self, target_sheet, startRow = 0, endRow=None, startCol=0, endCol=None):

        if target_sheet not in self.get_sheet_names():
            return False

        sht = self.wb.sheet_by_name(target_sheet)

        if not endRow:
            endRow = sht.nrows
            endRowText = "END"

        if not endCol:
            endColText = "END"

        data = list()

        #logging.debug("Looking for data, rows {} to {}, columns {} to {}".format(startRow, endRowText, startCol, endColText))

        for rowx in range(startRow,endRow):
            thisRow = sht.row_values(rowx, startCol, endCol)
            data.append(thisRow)

        #print(data)
        if len(data):
            logging.debug("Got data table from {}, {} rows, {} columns".format(target_sheet,len(data),len(data[0])))
        else:
            logging.debug("No data found in {}".format(target_sheet))

        return data
    
    def get_row_as_dict(self, table_data, rownum, num_cols = -1, flg_length_mismatch = True):
        header_row = [item for item in table_data[0] if item]
        data_row = [item for item in table_data[rownum] if item != ""]
        
        if num_cols != -1:
            header_row = header_row[0:num_cols]
            data_row = data_row[0:num_cols]

        if flg_length_mismatch:
            #print(header_row)
            #print(data_row)
            assert(len(header_row) == len(data_row)), "Mismatch header length {}, row length {}".format(len(header_row),len(data_row))
        
        return dict(list(zip(header_row, data_row)))
    
    def get_table_all(self, target_sheet, dataType="str"):
        
        sheet = self.wb.sheet_by_name(target_sheet)

        data = list()
        for i in range(sheet.nrows):
            if dataType == "str":
                data.append(sheet.row_values(i)) #drop all the values in the rows into data
            else:
                raise Exception("Unsupported data type")

        logging.debug("Got data table from {}, {} rows, {} columns".format(target_sheet,len(data),len(data[0])))

        return data

class ExcelBookRead(object):
    """
    import xlrd
    book = xlrd.open_workbook("myfile.xls")
    print "The number of worksheets is", book.nsheets
    print "Worksheet name(s):", book.sheet_names()
    sh = book.sheet_by_index(0)
    print sh.name, sh.nrows, sh.ncols
    print "Cell D30 is", sh.cell_value(rowx=29, colx=3)
    for rx in range(sh.nrows):
        print sh.row(rx)
    # Refer to docs for more details.
    # Feedback on API is welcomed.
    """


    def __init__(self, excelPath):
        self.excelPath = excelPath
        logging.debug("Excel file at: {}".format(self.excelPath))

    def get_sheet_names(self):
        wb = xlrd.open_workbook(self.excelPath)

        sheetNames = wb.sheet_names()
        logging.debug("Found {} sheet names: {}".format(len(sheetNames),sheetNames))
        return sheetNames

    def get_table_literal(self, targetSheet, startRow = 0, endRow=None, startCol=0, endCol=None):

        wb = xlrd.open_workbook(self.excelPath)

        if targetSheet not in self.get_sheet_names():
            return False

        sht = wb.sheet_by_name(targetSheet)

        if not endRow:
            endRow = sht.nrows
            endRowText = "END"

        if not endCol:
            endColText = "END"

        data = list()

        #logging.debug("Looking for data, rows {} to {}, columns {} to {}".format(startRow, endRowText, startCol, endColText))

        for rowx in range(startRow,endRow):
            thisRow = sht.row_values(rowx, startCol, endCol)
            data.append(thisRow)

        #print(data)
        if len(data):
            logging.debug("Got data table from {}, {} rows, {} columns".format(targetSheet,len(data),len(data[0])))
        else:
            logging.debug("No data found in {}".format(targetSheet))

        return data

    def get_table_all(self, targetSheet, dataType="str"):

        wb = xlrd.open_workbook(self.excelPath)
        sheet = wb.sheet_by_name(targetSheet)

        data = list()
        for i in range(sheet.nrows):
            if dataType == "str":
                data.append(sheet.row_values(i)) #drop all the values in the rows into data
            else:
                raise Exception("Unsupported data type")

        logging.debug("Got data table from {}, {} rows, {} columns".format(targetSheet,len(data),len(data[0])))

        return data
